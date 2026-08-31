import json

from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import (
    Table,
    TableStyleInfo,
)

from app.config.settings import settings
from app.domain.models.batch import Batch
from app.domain.models.document_evidence import (
    DocumentEvidence,
    EvidenceSeverity,
)
from app.domain.services.batch_cross_validation_service import (
    BatchCrossValidationService,
)
from app.infrastructure.repositories.analysis_memory_repository import (
    AnalysisMemoryRepository,
)
from app.domain.services.batch_finding_aggregation_service import (
    BatchFindingAggregationService,
)
from app.frontend.exports.batch_export_view_builder import (
    BatchExportViewBuilder,
)
from app.frontend.investigations.builders.investigation_view_builder import (
    InvestigationViewBuilder,
)
from app.frontend.investigations.services.investigation_status_resolver import (
    InvestigationStatusResolver,
)
from app.frontend.view_models.analysis_view_builder import (
    AnalysisViewBuilder,
)


class BatchExcelExportService:

    SUMMARY_SHEET_NAME = "00 - Resumo"
    DOCUMENTS_SHEET_NAME = "01 - Documentos"
    FINDINGS_SHEET_NAME = "02 - Achados por tipo"
    EVIDENCES_SHEET_NAME = "03 - Evidências"
    FINANCIAL_SHEET_NAME = "04 - Financeiro"
    PROMPT_INJECTION_SHEET_NAME = "05 - Prompt Injection"
    CONCEALMENT_SHEET_NAME = "06 - Ocultação visual"
    LOCATIONS_SHEET_NAME = "07 - Localizações"
    COMPARISONS_SHEET_NAME = "08 - Comparações"
    TECHNICAL_DATA_SHEET_NAME = "09 - Dados técnicos"
    IMAGES_SHEET_NAME = "10 - Imagens"

    def __init__(self) -> None:
        self._analysis_repository = (
            AnalysisMemoryRepository()
        )

        self._cross_validation_service = (
            BatchCrossValidationService()
        )

        self._finding_aggregation_service = (
            BatchFindingAggregationService()
        )

        self._analysis_view_builder = (
            AnalysisViewBuilder()
        )

        self._investigation_view_builder = (
            InvestigationViewBuilder()
        )

        self._status_resolver = (
            InvestigationStatusResolver()
        )

        self._export_view_builder = (
            BatchExportViewBuilder()
        )

    def export(
        self,
        batch: Batch,
    ) -> dict[str, Any]:
        output_dir = (
            settings.REPORTS_DIR
            / "batches"
        )

        output_dir.mkdir(
            parents=True,
            exist_ok=True,
        )

        filename = (
            f"docdna_lote_"
            f"{str(batch.id)[:8]}"
            f".xlsx"
        )

        file_path = (
            output_dir
            / filename
        )

        export_view = (
            self._build_export_view(
                batch
            )
        )

        workbook = Workbook()

        summary_sheet = workbook.active
        summary_sheet.title = (
            self.SUMMARY_SHEET_NAME
        )

        documents_sheet = workbook.create_sheet(
            self.DOCUMENTS_SHEET_NAME
        )
        findings_sheet = workbook.create_sheet(
            self.FINDINGS_SHEET_NAME
        )
        evidences_sheet = workbook.create_sheet(
            self.EVIDENCES_SHEET_NAME
        )
        financial_sheet = workbook.create_sheet(
            self.FINANCIAL_SHEET_NAME
        )
        prompt_injection_sheet = workbook.create_sheet(
            self.PROMPT_INJECTION_SHEET_NAME
        )
        concealment_sheet = workbook.create_sheet(
            self.CONCEALMENT_SHEET_NAME
        )
        locations_sheet = workbook.create_sheet(
            self.LOCATIONS_SHEET_NAME
        )
        comparisons_sheet = workbook.create_sheet(
            self.COMPARISONS_SHEET_NAME
        )
        technical_data_sheet = workbook.create_sheet(
            self.TECHNICAL_DATA_SHEET_NAME
        )
        images_sheet = workbook.create_sheet(
            self.IMAGES_SHEET_NAME
        )

        exported_documents = (
            self._build_documents_sheet(
                worksheet=documents_sheet,
                batch=batch,
                export_view=export_view,
            )
        )

        exported_findings = (
            self._build_findings_sheet(
                worksheet=findings_sheet,
                export_view=export_view,
            )
        )

        exported_evidences = (
            self._build_evidences_sheet(
                worksheet=evidences_sheet,
                batch=batch,
            )
        )

        exported_financial = (
            self._build_financial_sheet(
                worksheet=financial_sheet,
                export_view=export_view,
            )
        )

        exported_prompt_injection = (
            self._build_prompt_injection_sheet(
                worksheet=prompt_injection_sheet,
                export_view=export_view,
            )
        )

        exported_concealment = (
            self._build_concealment_sheet(
                worksheet=concealment_sheet,
                export_view=export_view,
            )
        )

        exported_locations = (
            self._build_locations_sheet(
                worksheet=locations_sheet,
                export_view=export_view,
            )
        )

        exported_technical_data = (
            self._build_technical_data_sheet(
                worksheet=technical_data_sheet,
                export_view=export_view,
            )
        )

        exported_images = (
            self._build_images_sheet(
                worksheet=images_sheet,
                export_view=export_view,
            )
        )

        evidence_report = (
            self._cross_validation_service
            .build_evidence_report(
                batch
            )
        )

        exported_comparisons = (
            self._build_comparisons_sheet(
                worksheet=comparisons_sheet,
                evidences=(
                    evidence_report.evidences
                ),
                total_documents=(
                    batch.result.total_documents
                ),
            )
        )

        self._build_summary_sheet(
            worksheet=summary_sheet,
            export_view=export_view,
            exported_documents=exported_documents,
            exported_findings=exported_findings,
            exported_evidences=exported_evidences,
            exported_financial=exported_financial,
            exported_prompt_injection=(
                exported_prompt_injection
            ),
            exported_concealment=(
                exported_concealment
            ),
            exported_locations=exported_locations,
            exported_comparisons=exported_comparisons,
            exported_technical_data=(
                exported_technical_data
            ),
            exported_images=exported_images,
            highest_severity=(
                evidence_report.highest_severity
            ),
        )

        workbook.save(
            file_path
        )

        return {
            "batch_id": str(batch.id),
            "filename": filename,
            "file_path": str(file_path),
            "download_url": (
                f"/reports/batches/{filename}"
            ),
            "total_documents": (
                batch.result.total_documents
            ),
            "exported_documents": exported_documents,
            "exported_findings": exported_findings,
            "exported_evidences": exported_evidences,
            "exported_financial": exported_financial,
            "exported_prompt_injection": (
                exported_prompt_injection
            ),
            "exported_concealment": (
                exported_concealment
            ),
            "exported_locations": exported_locations,
            "exported_comparisons": exported_comparisons,
            "exported_technical_data": (
                exported_technical_data
            ),
            "exported_images": exported_images,
            "message": (
                "Relatório do lote "
                "exportado com sucesso."
            ),
        }

    def _build_export_view(
        self,
        batch: Batch,
    ) -> dict[str, Any]:
        analyses: list[dict[str, Any]] = []
        analysis_views: list[dict[str, Any]] = []
        document_analytical_statuses = {}

        for batch_document in batch.documents:
            analysis_id = batch_document.analysis_id

            if analysis_id is None:
                continue

            analysis = self._get_analysis(
                analysis_id
            )

            if analysis is None:
                continue

            analyses.append(analysis)

            analysis_view = (
                self._analysis_view_builder
                .build(analysis)
            )

            analysis_views.append(
                analysis_view
            )

            investigation_cards = (
                self
                ._investigation_view_builder
                .build_cards(
                    analysis_id=analysis_id,
                    analysis_view=analysis_view,
                )
            )

            analytical_status = (
                self._status_resolver
                .resolve(
                    investigation_cards
                )
            )

            document_analytical_statuses[
                str(analysis_id)
            ] = analytical_status

        finding_summaries = (
            self
            ._finding_aggregation_service
            .aggregate(
                analyses
            )
        )

        return (
            self._export_view_builder
            .build(
                batch=batch,
                finding_summaries=(
                    finding_summaries
                ),
                document_analytical_statuses=(
                    document_analytical_statuses
                ),
                analysis_views=analysis_views,
            )
        )

    def _build_summary_sheet(
        self,
        worksheet,
        export_view: dict[str, Any],
        exported_documents: int,
        exported_findings: int,
        exported_evidences: int,
        exported_financial: int,
        exported_prompt_injection: int,
        exported_concealment: int,
        exported_locations: int,
        exported_comparisons: int,
        exported_technical_data: int,
        exported_images: int,
        highest_severity: EvidenceSeverity | None,
    ) -> None:
        summary = (
            export_view[
                "summary"
            ]
        )

        worksheet.append(
            [
                "Campo",
                "Valor",
            ]
        )

        rows = [
            (
                "ID do lote",
                summary["batch_id"],
            ),
            (
                "Tipo da análise",
                (
                    "Análise de documento individual"
                    if summary["total_documents"] == 1
                    else "Análise em lote"
                ),
            ),
            (
                "Observação",
                (
                    "Esta exportação corresponde à análise de um único "
                    "documento. Indicadores como 1/1 e 100% representam "
                    "a presença do achado neste documento e não um padrão "
                    "comparativo entre documentos."
                    if summary["total_documents"] == 1
                    else (
                        "Esta exportação consolida os resultados dos "
                        "documentos analisados no lote."
                    )
                ),
            ),
            (
                "Status do processamento",
                summary[
                    "processing_status_label"
                ],
            ),
            (
                "Situação analítica",
                summary[
                    "analytical_status_label"
                ],
            ),
            (
                "Criado em",
                summary["created_at"],
            ),
            (
                "Iniciado em",
                summary["started_at"],
            ),
            (
                "Finalizado em",
                summary["finished_at"],
            ),
            (
                "Total de documentos",
                summary[
                    "total_documents"
                ],
            ),
            (
                "Documentos concluídos",
                summary[
                    "completed_documents"
                ],
            ),
            (
                "Documentos com falha",
                summary[
                    "failed_documents"
                ],
            ),
            (
                "Documentos pendentes",
                summary[
                    "pending_documents"
                ],
            ),
            (
                "Documentos em processamento",
                summary[
                    "processing_documents"
                ],
            ),
            (
                "Progresso",
                summary[
                    "progress_label"
                ],
            ),
            (
                "Alta prioridade",
                summary[
                    "alert_documents"
                ],
            ),
            (
                "Revisão recomendada",
                summary[
                    "attention_documents"
                ],
            ),
            (
                "Sem apontamentos",
                summary[
                    "clear_documents"
                ],
            ),
            (
                "Análise incompleta",
                summary[
                    "not_executed_documents"
                ],
            ),
            (
                "Documentos classificados",
                summary[
                    "classified_documents"
                ],
            ),
            (
                "Tipos de achado",
                summary[
                    "finding_type_count"
                ],
            ),
            (
                "Documentos exportados",
                exported_documents,
            ),
            (
                "Achados agregados exportados",
                exported_findings,
            ),
            (
                "Evidências individuais exportadas",
                exported_evidences,
            ),
            (
                "Registros financeiros exportados",
                exported_financial,
            ),
            (
                "Evidências de Prompt Injection exportadas",
                exported_prompt_injection,
            ),
            (
                "Achados de ocultação visual exportados",
                exported_concealment,
            ),
            (
                "Localizações visuais exportadas",
                exported_locations,
            ),
            (
                "Comparações do lote exportadas",
                exported_comparisons,
            ),
            (
                "Registros técnicos exportados",
                exported_technical_data,
            ),
            (
                "Fingerprints de imagem exportados",
                exported_images,
            ),
            (
                "Severidade comparativa predominante",
                (
                    self._translate_severity(
                        highest_severity
                    )
                    if highest_severity
                    else "Sem apontamentos"
                ),
            ),
        ]

        for row in rows:
            worksheet.append(
                row
            )

        self._style_worksheet(
            worksheet=worksheet,
            freeze_panes="A2",
            auto_filter=False,
        )

        worksheet.column_dimensions[
            "A"
        ].width = 42

        worksheet.column_dimensions[
            "B"
        ].width = 52

        self._style_summary_statuses(
            worksheet=worksheet,
            analytical_status=(
                summary[
                    "analytical_status"
                ]
            ),
        )

    def _build_documents_sheet(
        self,
        worksheet,
        batch: Batch,
        export_view: dict[str, Any],
    ) -> int:
        documents_by_analysis_id = {
            str(document["analysis_id"]): document
            for document in export_view["documents"]
            if document["analysis_id"]
        }

        headers = [
            "Documento",
            "Status no lote",
            "Situação analítica",
            "Código analítico",
            "ID da análise",
            "Tamanho em bytes",
            "SHA-256",
            "Páginas",
            "Título do PDF",
            "Autor do PDF",
            "Produtor do PDF",
            "Data de criação",
            "Data de modificação",
            "ITF",
            "QR Code",
            "Code 39",
            "Outros códigos",
            "Sequências numéricas",
            "Fontes das sequências",
            "Validações estruturais",
            "Comparações linha × código",
            "Total de evidências",
            "Status da localização visual",
            "Erro de processamento",
        ]

        worksheet.append(headers)

        exported_documents = 0

        for batch_document in batch.documents:
            analysis_id_key = (
                str(batch_document.analysis_id)
                if batch_document.analysis_id
                else ""
            )

            exported_document = documents_by_analysis_id.get(
                analysis_id_key
            )

            analytical_status = (
                exported_document["analytical_status"]
                if exported_document
                else "not_executed"
            )

            analytical_status_label = (
                exported_document["analytical_status_label"]
                if exported_document
                else "Análise incompleta"
            )

            analysis = self._get_analysis(
                batch_document.analysis_id
            )

            if analysis is None:
                worksheet.append(
                    [
                        batch_document.original_filename,
                        self._translate_document_status(
                            batch_document.status.value
                        ),
                        analytical_status_label,
                        analytical_status,
                        self._optional_uuid(
                            batch_document.analysis_id
                        ),
                        None,
                        None,
                        None,
                        None,
                        None,
                        None,
                        None,
                        None,
                        None,
                        None,
                        None,
                        None,
                        None,
                        None,
                        None,
                        None,
                        0,
                        None,
                        batch_document.error_message,
                    ]
                )
                continue

            barcodes = analysis.get(
                "barcodes",
                [],
            )

            barcode_groups = self._group_barcodes(
                barcodes
            )

            printed_numeric_lines = analysis.get(
                "printed_numeric_lines",
                [],
            )

            numeric_line_validations = analysis.get(
                "numeric_line_validations",
                [],
            )

            barcode_line_comparisons = analysis.get(
                "barcode_line_comparisons",
                [],
            )

            numeric_line_locations = analysis.get(
                "numeric_line_locations",
                [],
            )

            evidences = analysis.get(
                "evidences",
                [],
            )

            pdf_info = analysis.get(
                "pdf_info"
            )

            worksheet.append(
                [
                    batch_document.original_filename,
                    self._translate_document_status(
                        batch_document.status.value
                    ),
                    analytical_status_label,
                    analytical_status,
                    self._optional_uuid(
                        batch_document.analysis_id
                    ),
                    analysis.get("size_bytes"),
                    analysis.get("sha256"),
                    self._safe_attribute(
                        pdf_info,
                        "page_count",
                    ),
                    self._safe_attribute(
                        pdf_info,
                        "title",
                    ),
                    self._safe_attribute(
                        pdf_info,
                        "author",
                    ),
                    self._safe_attribute(
                        pdf_info,
                        "producer",
                    ),
                    self._safe_attribute(
                        pdf_info,
                        "creation_date",
                    ),
                    self._safe_attribute(
                        pdf_info,
                        "modification_date",
                    ),
                    self._join_values(
                        barcode_groups["itf"]
                    ),
                    self._join_values(
                        barcode_groups["qr_code"]
                    ),
                    self._join_values(
                        barcode_groups["code_39"]
                    ),
                    self._join_values(
                        barcode_groups["others"]
                    ),
                    self._format_numeric_lines(
                        printed_numeric_lines
                    ),
                    self._format_numeric_line_sources(
                        printed_numeric_lines
                    ),
                    self._format_validations(
                        numeric_line_validations
                    ),
                    self._format_comparisons(
                        barcode_line_comparisons
                    ),
                    len(evidences),
                    self._format_locations(
                        numeric_line_locations
                    ),
                    batch_document.error_message,
                ]
            )

            exported_documents += 1

        self._style_worksheet(
            worksheet=worksheet,
            freeze_panes="A2",
            auto_filter=False,
        )

        self._set_documents_column_widths(
            worksheet
        )

        self._style_document_statuses(
            worksheet
        )

        self._add_excel_table(
            worksheet=worksheet,
            table_name="DocDNA_Documentos",
        )

        return exported_documents

    def _build_findings_sheet(
            self,
            worksheet,
            export_view: dict[str, Any],
    ) -> int:
        headers = [
            "Código",
            "Tipo de achado",
            "Documentos afetados",
            "Total de documentos",
            "Prevalência",
            "Ocorrências",
            "Maior confiança",
            "IDs das análises afetadas",
        ]

        worksheet.append(
            headers
        )

        findings = (
            export_view[
                "findings_by_type"
            ]
        )

        if not findings:
            worksheet.append(
                [
                    "SEM_ACHADOS",
                    (
                        "Nenhum achado "
                        "agregado identificado"
                    ),
                    0,
                    export_view[
                        "summary"
                    ][
                        "total_documents"
                    ],
                    0.0,
                    0,
                    0.0,
                    "",
                ]
            )
        else:
            for finding in findings:
                worksheet.append(
                    [
                        finding["code"],
                        finding["title"],
                        finding[
                            "affected_documents"
                        ],
                        finding[
                            "total_documents"
                        ],
                        finding[
                            "prevalence_ratio"
                        ],
                        finding[
                            "occurrence_count"
                        ],
                        finding[
                            "highest_confidence"
                        ],
                        self._join_values(
                            finding[
                                "affected_document_ids"
                            ]
                        ),
                    ]
                )

        self._style_worksheet(
            worksheet=worksheet,
            freeze_panes="A2",
            auto_filter=False,
        )

        for cell in worksheet["E"][1:]:
            cell.number_format = (
                "0.0%"
            )

        for cell in worksheet["G"][1:]:
            cell.number_format = (
                "0%"
            )

        widths = {
            "A": 32,
            "B": 38,
            "C": 22,
            "D": 22,
            "E": 18,
            "F": 18,
            "G": 18,
            "H": 70,
        }

        for column, width in (
                widths.items()
        ):
            worksheet.column_dimensions[
                column
            ].width = width

        self._add_excel_table(
            worksheet=worksheet,
            table_name=(
                "DocDNA_AchadosPorTipo"
            ),
        )

        return len(findings)

    def _build_evidences_sheet(
        self,
        worksheet,
        batch: Batch,
    ) -> int:
        headers = [
            "Documento",
            "ID da análise",
            "Código da evidência",
            "Título",
            "Descrição",
            "Severidade",
            "Detector",
            "Confiança",
        ]

        worksheet.append(headers)

        exported_evidences = 0

        for batch_document in batch.documents:
            analysis = self._get_analysis(
                batch_document.analysis_id
            )

            if analysis is None:
                continue

            evidences = analysis.get(
                "evidences",
                [],
            )

            for evidence in evidences:
                severity = self._enum_value(
                    self._safe_attribute(
                        evidence,
                        "severity",
                    )
                )

                confidence = self._safe_attribute(
                    evidence,
                    "confidence",
                )

                worksheet.append(
                    [
                        batch_document.original_filename,
                        self._optional_uuid(
                            batch_document.analysis_id
                        ),
                        self._safe_attribute(
                            evidence,
                            "code",
                        ),
                        self._safe_attribute(
                            evidence,
                            "title",
                        ),
                        self._safe_attribute(
                            evidence,
                            "description",
                        ),
                        self._translate_severity_value(
                            severity
                        ),
                        (
                            self._safe_attribute(
                                evidence,
                                "detector",
                            )
                            or self._safe_attribute(
                                evidence,
                                "source",
                            )
                        ),
                        self._format_confidence(
                            confidence
                        ),
                    ]
                )

                exported_evidences += 1

        if exported_evidences == 0:
            worksheet.append(
                [
                    "Nenhuma evidência genérica foi identificada.",
                ]
                + [None] * (len(headers) - 1)
            )

        self._style_worksheet(
            worksheet=worksheet,
            freeze_panes="A2",
            auto_filter=False,
        )

        column_widths = {
            "A": 38,
            "B": 38,
            "C": 30,
            "D": 42,
            "E": 70,
            "F": 18,
            "G": 44,
            "H": 16,
        }

        for column, width in column_widths.items():
            worksheet.column_dimensions[
                column
            ].width = width

        self._add_excel_table(
            worksheet=worksheet,
            table_name="DocDNA_Evidencias",
        )

        return exported_evidences

    def _build_financial_sheet(
        self,
        worksheet,
        export_view: dict[str, Any],
    ) -> int:
        headers = [
            "Documento",
            "ID da análise",
            "Tipo de registro",
            "Índice da linha",
            "Índice do barcode",
            "Linha digitável",
            "Tipo da linha",
            "Status",
            "Método de validação",
            "DVs válidos",
            "Total de DVs",
            "Barcode calculado",
            "Barcode detectado",
            "Mensagem",
        ]

        worksheet.append(headers)

        records = export_view.get(
            "financial",
            [],
        )

        if not records:
            worksheet.append(
                [
                    "Nenhum apontamento financeiro foi identificado.",
                ]
                + [None] * (len(headers) - 1)
            )

        for record in records:
            worksheet.append(
                [
                    record.get("filename"),
                    record.get("analysis_id"),
                    record.get("record_type_label"),
                    record.get("line_index"),
                    record.get("barcode_index"),
                    record.get("numeric_line"),
                    record.get("line_type_label"),
                    record.get("status_label"),
                    record.get("validation_method_label"),
                    record.get("valid_check_digits"),
                    record.get("total_check_digits"),
                    record.get("converted_barcode"),
                    record.get("detected_barcode"),
                    record.get("message"),
                ]
            )

        self._style_worksheet(
            worksheet=worksheet,
            freeze_panes="A2",
            auto_filter=False,
        )

        self._set_column_widths(
            worksheet,
            [
                38, 38, 24, 16, 18, 70, 28,
                28, 24, 14, 14, 62, 62, 70,
            ],
        )

        self._add_excel_table(
            worksheet=worksheet,
            table_name="DocDNA_Financeiro",
        )

        return len(records)

    def _build_prompt_injection_sheet(
        self,
        worksheet,
        export_view: dict[str, Any],
    ) -> int:
        headers = [
            "Documento",
            "ID da análise",
            "Índice da evidência",
            "Risco",
            "Score da análise",
            "Código",
            "Categoria",
            "Detector",
            "Página",
            "Origem",
            "Confiança",
            "Peso",
            "Score ponderado",
            "Regra identificada",
            "Trecho original",
            "Trecho normalizado",
            "Idioma",
            "Fonte",
            "Tamanho da fonte",
            "Cor da fonte",
            "Maior fonte da página",
            "Método de análise",
            "Grupos de sinais",
            "Sinais identificados",
            "Descrição",
        ]

        worksheet.append(headers)

        records = export_view.get(
            "prompt_injection",
            [],
        )

        if not records:
            worksheet.append(
                [
                    "Nenhum apontamento de Prompt Injection foi identificado.",
                ]
                + [None] * (len(headers) - 1)
            )

        for record in records:
            worksheet.append(
                [
                    record.get("filename"),
                    record.get("analysis_id"),
                    record.get("evidence_index"),
                    record.get("risk_label"),
                    self._ratio_value(record.get("score")),
                    record.get("code"),
                    record.get("category_label"),
                    record.get("detector"),
                    record.get("page_number"),
                    record.get("source_label"),
                    self._ratio_value(record.get("confidence")),
                    self._ratio_value(record.get("weight")),
                    self._ratio_value(record.get("weighted_score")),
                    record.get("matched_rule"),
                    record.get("original_excerpt"),
                    record.get("normalized_excerpt"),
                    record.get("language"),
                    record.get("font_name"),
                    record.get("font_size"),
                    record.get("font_color"),
                    record.get("maximum_font_size"),
                    record.get("analysis_method"),
                    self._join_values(
                        record.get("signal_groups", [])
                    ),
                    self._format_json_value(
                        record.get("matched_signals")
                    ),
                    record.get("description"),
                ]
            )

        for column in ("E", "K", "L", "M"):
            for cell in worksheet[column][1:]:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "0.0%"

        self._style_worksheet(
            worksheet=worksheet,
            freeze_panes="A2",
            auto_filter=False,
        )

        self._set_column_widths(
            worksheet,
            [
                38, 38, 18, 18, 18, 34, 32, 42,
                12, 24, 16, 14, 18, 48, 80, 80,
                14, 28, 18, 20, 20, 24, 40, 58, 70,
            ],
        )

        self._add_excel_table(
            worksheet=worksheet,
            table_name="DocDNA_PromptInjection",
        )

        return len(records)

    def _build_concealment_sheet(
        self,
        worksheet,
        export_view: dict[str, Any],
    ) -> int:
        headers = [
            "Documento",
            "ID da análise",
            "Índice do finding",
            "Tipo",
            "Código",
            "Detector",
            "Página",
            "Texto",
            "Descrição",
            "Fonte",
            "Tamanho da fonte",
            "Cor da fonte",
            "Confiança",
            "Sinais técnicos",
            "Texto branco/quase branco",
            "Texto pequeno",
            "Pequeno relativo à página",
            "Conteúdo instrucional",
            "Coordenadas",
        ]

        worksheet.append(headers)

        records = export_view.get(
            "concealment",
            [],
        )

        if not records:
            worksheet.append(
                [
                    "Nenhum achado de ocultação visual foi identificado.",
                ]
                + [None] * (len(headers) - 1)
            )

        for record in records:
            signal_labels = record.get(
                "signal_labels",
                [],
            )

            if not signal_labels:
                signal_labels = record.get(
                    "signals",
                    [],
                )

            worksheet.append(
                [
                    record.get("filename"),
                    record.get("analysis_id"),
                    record.get("finding_index"),
                    record.get("concealment_type_label"),
                    record.get("code"),
                    record.get("detector"),
                    record.get("page_number"),
                    record.get("text"),
                    record.get("description"),
                    record.get("font_name"),
                    record.get("font_size"),
                    record.get("font_color_hex"),
                    self._ratio_value(record.get("confidence")),
                    self._join_values(signal_labels),
                    self._yes_no(record.get("is_near_white")),
                    self._yes_no(record.get("is_small_text")),
                    self._yes_no(
                        record.get("is_relative_small_text")
                    ),
                    self._yes_no(
                        record.get("is_instruction_like")
                    ),
                    record.get("coordinates_label"),
                ]
            )

        for cell in worksheet["M"][1:]:
            if isinstance(cell.value, (int, float)):
                cell.number_format = "0.0%"

        self._style_worksheet(
            worksheet=worksheet,
            freeze_panes="A2",
            auto_filter=False,
        )

        self._set_column_widths(
            worksheet,
            [
                38, 38, 18, 32, 32, 40, 12, 80, 70,
                28, 18, 20, 16, 55, 24, 18, 24, 24, 48,
            ],
        )

        self._add_excel_table(
            worksheet=worksheet,
            table_name="DocDNA_OcultacaoVisual",
        )

        return len(records)

    def _build_locations_sheet(
        self,
        worksheet,
        export_view: dict[str, Any],
    ) -> int:
        headers = [
            "Documento",
            "ID da análise",
            "Tipo de localização",
            "Índice de referência",
            "Código de referência",
            "Detector",
            "Página",
            "Conteúdo localizado",
            "X",
            "Y",
            "Largura",
            "Altura",
            "Confiança",
            "Fonte",
            "Tamanho da fonte",
            "Cor da fonte",
            "Coordenadas",
            "Localizado",
            "Mensagem",
            "Imagem de origem",
            "Imagem anotada",
        ]

        worksheet.append(headers)

        records = export_view.get(
            "locations",
            [],
        )

        if not records:
            worksheet.append(
                [
                    "Nenhuma localização visual foi produzida para este lote.",
                ]
                + [None] * (len(headers) - 1)
            )

        for record in records:
            worksheet.append(
                [
                    record.get("filename"),
                    record.get("analysis_id"),
                    record.get("location_type_label"),
                    record.get("reference_index"),
                    record.get("reference_code"),
                    record.get("detector"),
                    record.get("page_number"),
                    record.get("matched_content"),
                    record.get("left"),
                    record.get("top"),
                    record.get("width"),
                    record.get("height"),
                    self._ratio_value(record.get("confidence")),
                    record.get("font_name"),
                    record.get("font_size"),
                    record.get("font_color_hex"),
                    record.get("coordinates_label"),
                    self._yes_no(record.get("located")),
                    record.get("message"),
                    record.get("source_image_url"),
                    record.get("annotated_image_url"),
                ]
            )

            row_index = worksheet.max_row
            self._set_hyperlink(
                worksheet.cell(
                    row=row_index,
                    column=20,
                )
            )
            self._set_hyperlink(
                worksheet.cell(
                    row=row_index,
                    column=21,
                )
            )

        for cell in worksheet["M"][1:]:
            if isinstance(cell.value, (int, float)):
                cell.number_format = "0.0%"

        self._style_worksheet(
            worksheet=worksheet,
            freeze_panes="A2",
            auto_filter=False,
        )

        self._set_column_widths(
            worksheet,
            [
                38, 38, 28, 18, 34, 42, 12, 80,
                14, 14, 14, 14, 16, 28, 18, 20,
                48, 14, 60, 62, 62,
            ],
        )

        self._add_excel_table(
            worksheet=worksheet,
            table_name="DocDNA_Localizacoes",
        )

        return len(records)

    def _build_technical_data_sheet(
        self,
        worksheet,
        export_view: dict[str, Any],
    ) -> int:
        headers = [
            "Documento",
            "ID da análise",
            "Enviado em",
            "Tamanho em bytes",
            "Tamanho formatado",
            "SHA-256",
            "Páginas",
            "Título do PDF",
            "Autor do PDF",
            "Criador do PDF",
            "Produtor do PDF",
            "Data de criação",
            "Data de modificação",
            "Versão do PDF",
            "Possui texto nativo",
            "Caracteres de texto nativo",
            "Páginas com texto nativo",
            "Caracteres OCR",
            "Páginas processadas por OCR",
            "Páginas com texto OCR",
            "Idioma OCR",
            "Documento normalizado",
            "Páginas normalizadas",
            "Text spans normalizados",
            "Palavras normalizadas",
            "Caracteres normalizados",
            "Caracteres após normalização",
            "Páginas normalizadas com texto",
            "Imagens extraídas",
            "Fingerprints de imagem",
            "Códigos de barras",
            "Formatos de código",
            "Páginas com código",
            "Sequências numéricas",
            "Fontes das sequências",
            "Dígitos em sequências",
            "Linhas válidas",
            "Linhas inválidas",
            "Linhas inconclusivas",
            "Comparações compatíveis",
            "Comparações divergentes",
            "Comparações inconclusivas",
            "Total de evidências",
        ]

        worksheet.append(headers)

        records = export_view.get(
            "technical_data",
            [],
        )

        if not records:
            worksheet.append(
                [
                    "Nenhum dado técnico foi disponibilizado para exportação.",
                ]
                + [None] * (len(headers) - 1)
            )

        for record in records:
            worksheet.append(
                [
                    record.get("filename"),
                    record.get("analysis_id"),
                    record.get("uploaded_at"),
                    record.get("size_bytes"),
                    record.get("formatted_size"),
                    record.get("sha256"),
                    record.get("page_count"),
                    record.get("pdf_title"),
                    record.get("pdf_author"),
                    record.get("pdf_creator"),
                    record.get("pdf_producer"),
                    record.get("pdf_creation_date"),
                    record.get("pdf_modification_date"),
                    record.get("pdf_version"),
                    self._yes_no(
                        record.get("has_native_text")
                    ),
                    record.get("native_text_character_count"),
                    record.get("native_text_pages", 0),
                    record.get("ocr_character_count"),
                    record.get("ocr_pages_processed"),
                    record.get("ocr_pages_with_text", 0),
                    record.get("ocr_language"),
                    self._yes_no(
                        record.get("has_normalized_document")
                    ),
                    record.get("normalized_document_page_count"),
                    record.get("normalized_document_text_span_count"),
                    record.get("normalized_document_word_count"),
                    record.get("normalized_document_character_count"),
                    record.get(
                        "normalized_document_normalized_character_count"
                    ),
                    record.get(
                        "normalized_document_pages_with_text"
                    ),
                    record.get("image_count"),
                    record.get("image_fingerprint_count"),
                    record.get("barcode_count"),
                    record.get("barcode_formats"),
                    record.get("barcode_pages"),
                    record.get("printed_numeric_line_count"),
                    record.get("printed_numeric_line_sources"),
                    record.get("printed_numeric_digit_total"),
                    record.get("valid_numeric_line_count"),
                    record.get("invalid_numeric_line_count"),
                    record.get("inconclusive_numeric_line_count"),
                    record.get("barcode_line_match_count"),
                    record.get("barcode_line_mismatch_count"),
                    record.get("barcode_line_inconclusive_count"),
                    record.get("evidence_count"),
                ]
            )

        self._style_worksheet(
            worksheet=worksheet,
            freeze_panes="A2",
            auto_filter=False,
        )

        self._set_column_widths(
            worksheet,
            [
                38, 38, 24, 18, 20, 68, 12, 36, 32, 32, 32,
                24, 24, 16, 20, 24, 28, 18, 24, 24, 16, 22,
                20, 22, 22, 24, 26, 26, 18, 22, 18, 28, 22,
                22, 28, 20, 16, 16, 18, 20, 20, 20, 18,
            ],
        )

        self._add_excel_table(
            worksheet=worksheet,
            table_name="DocDNA_DadosTecnicos",
        )

        return len(records)

    def _build_images_sheet(
        self,
        worksheet,
        export_view: dict[str, Any],
    ) -> int:
        headers = [
            "Documento",
            "ID da análise",
            "Imagens extraídas no documento",
            "Índice do fingerprint",
            "Página",
            "Largura",
            "Altura",
            "MIME type",
            "DPI",
            "Descrição",
            "Confiança",
            "SHA-256 da imagem",
            "Perceptual hash",
            "Average hash",
            "Difference hash",
            "Posição X",
            "Posição Y",
            "Largura da região",
            "Altura da região",
        ]

        worksheet.append(headers)

        records = export_view.get(
            "images",
            [],
        )

        if not records:
            worksheet.append(
                [
                    "Nenhum fingerprint de imagem foi produzido para este lote.",
                ]
                + [None] * (len(headers) - 1)
            )

        for record in records:
            worksheet.append(
                [
                    record.get("filename"),
                    record.get("analysis_id"),
                    record.get("total_extracted_images"),
                    record.get("fingerprint_index"),
                    record.get("page_number"),
                    record.get("width"),
                    record.get("height"),
                    record.get("mime_type"),
                    record.get("dpi"),
                    record.get("description"),
                    record.get("confidence"),
                    record.get("image_hash"),
                    record.get("perceptual_hash"),
                    record.get("average_hash"),
                    record.get("difference_hash"),
                    record.get("location_x"),
                    record.get("location_y"),
                    record.get("location_width"),
                    record.get("location_height"),
                ]
            )

        self._style_worksheet(
            worksheet=worksheet,
            freeze_panes="A2",
            auto_filter=False,
        )

        self._set_column_widths(
            worksheet,
            [
                38, 38, 26, 22, 12, 14, 14, 22, 16, 54,
                18, 68, 38, 38, 38, 16, 16, 20, 20,
            ],
        )

        self._add_excel_table(
            worksheet=worksheet,
            table_name="DocDNA_Imagens",
        )

        return len(records)

    def _build_comparisons_sheet(
        self,
        worksheet,
        evidences: list[DocumentEvidence],
        total_documents: int,
    ) -> int:
        headers = [
            "Código do apontamento",
            "Título",
            "Descrição",
            "Severidade",
            "Confiança",
            "Código ITF relacionado",
            "Quantidade de documentos",
            "Documentos envolvidos",
            "Sequências numéricas por documento",
            "Componente responsável",
        ]

        worksheet.append(headers)

        exported_comparisons = 0

        if not evidences:
            if total_documents == 1:
                message = (
                    "Comparação entre documentos não aplicável: "
                    "este lote contém apenas um documento. "
                    "São necessários dois ou mais documentos para "
                    "análises comparativas."
                )
            else:
                message = (
                    "Nenhum apontamento comparativo foi identificado "
                    "entre os documentos deste lote."
                )

            worksheet.append(
                [message]
                + [None] * (len(headers) - 1)
            )

        for evidence in evidences:
            metadata = dict(
                evidence.metadata
            )

            document_names = (
                self._extract_comparison_document_names(
                    metadata
                )
            )

            numeric_lines_by_document = (
                self._format_comparison_numeric_lines(
                    metadata
                )
            )

            worksheet.append(
                [
                    evidence.code,
                    evidence.title,
                    evidence.description,
                    self._translate_severity(
                        evidence.severity
                    ),
                    self._format_confidence(
                        evidence.confidence
                    ),
                    metadata.get("itf"),
                    (
                        metadata.get(
                            "document_count"
                        )
                        or len(
                            evidence.document_ids
                        )
                    ),
                    self._join_values(
                        document_names
                    ),
                    numeric_lines_by_document,
                    evidence.source,
                ]
            )

            exported_comparisons += 1

        self._style_worksheet(
            worksheet=worksheet,
            freeze_panes="A2",
            auto_filter=False,
        )

        column_widths = {
            "A": 42,
            "B": 52,
            "C": 80,
            "D": 18,
            "E": 16,
            "F": 60,
            "G": 24,
            "H": 54,
            "I": 80,
            "J": 52,
        }

        for column, width in column_widths.items():
            worksheet.column_dimensions[
                column
            ].width = width

        self._add_excel_table(
            worksheet=worksheet,
            table_name="DocDNA_Comparacoes",
        )

        return exported_comparisons

    def _extract_comparison_document_names(
        self,
        metadata: dict[str, Any],
    ) -> list[str]:
        document_names = metadata.get(
            "document_names",
            [],
        )

        if isinstance(document_names, list):
            normalized_names = [
                str(name).strip()
                for name in document_names
                if str(name).strip()
            ]

            if normalized_names:
                return normalized_names

        documents = metadata.get(
            "documents",
            [],
        )

        if not isinstance(documents, list):
            return []

        names: list[str] = []

        for document in documents:
            if not isinstance(document, dict):
                continue

            filename = str(
                document.get(
                    "filename",
                    "",
                )
            ).strip()

            if filename and filename not in names:
                names.append(filename)

        return names

    def _format_comparison_numeric_lines(
        self,
        metadata: dict[str, Any],
    ) -> str:
        documents = metadata.get(
            "documents",
            [],
        )

        if not isinstance(documents, list):
            return ""

        formatted_documents: list[str] = []

        for document in documents:
            if not isinstance(document, dict):
                continue

            filename = str(
                document.get(
                    "filename",
                    "Documento sem nome",
                )
            ).strip()

            numeric_lines = document.get(
                "numeric_lines",
                [],
            )

            if not isinstance(numeric_lines, list):
                numeric_lines = [
                    numeric_lines
                ]

            normalized_lines = self._join_values(
                numeric_lines
            )

            if not normalized_lines:
                normalized_lines = (
                    "Nenhuma sequência registrada"
                )

            formatted_documents.append(
                f"{filename}: {normalized_lines}"
            )

        return "\n\n".join(
            formatted_documents
        )

    def _get_analysis(
        self,
        analysis_id,
    ) -> dict[str, Any] | None:
        if analysis_id is None:
            return None

        return self._analysis_repository.get_by_id(
            analysis_id
        )

    def _group_barcodes(
        self,
        barcodes: list[Any],
    ) -> dict[str, list[str]]:
        grouped = {
            "itf": [],
            "qr_code": [],
            "code_39": [],
            "others": [],
        }

        for barcode in barcodes:
            content = str(
                self._safe_attribute(
                    barcode,
                    "content",
                )
                or ""
            ).strip()

            barcode_format = str(
                self._safe_attribute(
                    barcode,
                    "format",
                )
                or ""
            ).strip()

            normalized_format = (
                barcode_format
                .lower()
                .replace("-", "")
                .replace("_", "")
                .replace(" ", "")
            )

            if not content:
                continue

            if "itf" in normalized_format:
                grouped["itf"].append(
                    content
                )
                continue

            if (
                "qr" in normalized_format
                or "qrcode" in normalized_format
            ):
                grouped["qr_code"].append(
                    content
                )
                continue

            if (
                "code39" in normalized_format
                or normalized_format == "39"
            ):
                grouped["code_39"].append(
                    content
                )
                continue

            grouped["others"].append(
                f"{barcode_format}: {content}"
            )

        return grouped

    def _format_numeric_lines(
        self,
        lines: list[Any],
    ) -> str:
        values = [
            str(
                self._safe_attribute(
                    line,
                    "normalized_content",
                )
                or ""
            )
            for line in lines
        ]

        return self._join_values(values)

    def _format_numeric_line_sources(
        self,
        lines: list[Any],
    ) -> str:
        values = [
            str(
                self._safe_attribute(
                    line,
                    "source",
                )
                or ""
            )
            for line in lines
        ]

        return self._join_values(values)

    def _format_validations(
        self,
        validations: list[Any],
    ) -> str:
        formatted_values: list[str] = []

        for validation in validations:
            status = self._enum_value(
                self._safe_attribute(
                    validation,
                    "status",
                )
            )

            line_type = self._enum_value(
                self._safe_attribute(
                    validation,
                    "line_type",
                )
            )

            valid_digits = self._safe_attribute(
                validation,
                "valid_check_digits",
            )

            total_digits = self._safe_attribute(
                validation,
                "total_check_digits",
            )

            formatted_values.append(
                (
                    f"Sequência "
                    f"{self._safe_attribute(validation, 'line_index')}: "
                    f"{status}; "
                    f"tipo={line_type}; "
                    f"dígitos verificadores="
                    f"{valid_digits}/{total_digits}"
                )
            )

        return self._join_values(
            formatted_values
        )

    def _format_comparisons(
        self,
        comparisons: list[Any],
    ) -> str:
        formatted_values: list[str] = []

        for comparison in comparisons:
            status = self._enum_value(
                self._safe_attribute(
                    comparison,
                    "status",
                )
            )

            converted_barcode = self._safe_attribute(
                comparison,
                "converted_barcode",
            )

            detected_barcode = self._safe_attribute(
                comparison,
                "detected_barcode",
            )

            formatted_values.append(
                (
                    f"Sequência "
                    f"{self._safe_attribute(comparison, 'line_index')}: "
                    f"{status}; "
                    f"calculado={converted_barcode or '-'}; "
                    f"detectado={detected_barcode or '-'}"
                )
            )

        return self._join_values(
            formatted_values
        )

    def _format_locations(
        self,
        locations: list[Any],
    ) -> str:
        formatted_values: list[str] = []

        for location in locations:
            located = bool(
                self._safe_attribute(
                    location,
                    "located",
                )
            )

            if located:
                formatted_values.append(
                    (
                        f"Sequência "
                        f"{self._safe_attribute(location, 'line_index')}: "
                        f"localizada na página "
                        f"{self._safe_attribute(location, 'page_number')}"
                    )
                )
            else:
                formatted_values.append(
                    (
                        f"Sequência "
                        f"{self._safe_attribute(location, 'line_index')}: "
                        "não localizada"
                    )
                )

        return self._join_values(
            formatted_values
        )

    def _style_worksheet(
        self,
        worksheet,
        freeze_panes: str,
        auto_filter: bool,
    ) -> None:
        header_fill = PatternFill(
            fill_type="solid",
            fgColor="24184F",
        )

        header_font = Font(
            color="FFFFFF",
            bold=True,
        )

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

        worksheet.row_dimensions[1].height = 32

        for row in worksheet.iter_rows(
            min_row=2,
        ):
            for cell in row:
                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=True,
                )

        worksheet.freeze_panes = freeze_panes

        if auto_filter:
            worksheet.auto_filter.ref = (
                worksheet.dimensions
            )

        worksheet.sheet_view.showGridLines = False

    def _set_documents_column_widths(
        self,
        worksheet,
    ) -> None:
        widths = [
            38,  # Documento
            18,  # Status operacional
            24,  # Situação analítica
            18,  # Código analítico
            38,  # ID análise
            18,  # Tamanho
            68,  # SHA
            12,  # Páginas
            30,  # Título
            26,  # Autor
            30,  # Produtor
            24,  # Criação
            24,  # Modificação
            54,  # ITF
            70,  # QR
            34,  # Code39
            52,  # Outros
            70,  # Sequências
            24,  # Fontes
            70,  # Validações
            80,  # Comparações
            18,  # Evidências
            44,  # Localização
            46,  # Erro
        ]

        for index, width in enumerate(
            widths,
            start=1,
        ):
            column_letter = get_column_letter(
                index
            )

            worksheet.column_dimensions[
                column_letter
            ].width = width

    def _add_excel_table(
        self,
        *,
        worksheet,
        table_name: str,
    ) -> None:
        if worksheet.max_row < 2:
            return

        reference = (
            f"A1:"
            f"{get_column_letter(worksheet.max_column)}"
            f"{worksheet.max_row}"
        )

        table = Table(
            displayName=table_name,
            ref=reference,
        )

        style = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )

        table.tableStyleInfo = style

        worksheet.add_table(
            table
        )

    def _style_document_statuses(
        self,
        worksheet,
    ) -> None:
        fills = {
            "alert": "FCE8EC",
            "attention": "FFF4D6",
            "clear": "EAF6EE",
            "not_executed": "EEF2F6",
        }

        fonts = {
            "alert": "9F1239",
            "attention": "92400E",
            "clear": "166534",
            "not_executed": "475569",
        }

        for row_index in range(
            2,
            worksheet.max_row + 1,
        ):
            label_cell = worksheet.cell(
                row=row_index,
                column=3,
            )

            code_cell = worksheet.cell(
                row=row_index,
                column=4,
            )

            status = str(
                code_cell.value or ""
            ).strip()

            fill_color = fills.get(
                status
            )

            font_color = fonts.get(
                status
            )

            if fill_color:
                label_cell.fill = PatternFill(
                    fill_type="solid",
                    fgColor=fill_color,
                )

            if font_color:
                label_cell.font = Font(
                    color=font_color,
                    bold=True,
                )

    def _style_summary_statuses(
        self,
        *,
        worksheet,
        analytical_status: str,
    ) -> None:
        fills = {
            "alert": "FCE8EC",
            "attention": "FFF4D6",
            "clear": "EAF6EE",
            "not_executed": "EEF2F6",
        }

        fonts = {
            "alert": "9F1239",
            "attention": "92400E",
            "clear": "166534",
            "not_executed": "475569",
        }

        for row_index in range(
            2,
            worksheet.max_row + 1,
        ):
            label = worksheet.cell(
                row=row_index,
                column=1,
            ).value

            if label != "Situação analítica":
                continue

            value_cell = worksheet.cell(
                row=row_index,
                column=2,
            )

            fill_color = fills.get(
                analytical_status
            )

            font_color = fonts.get(
                analytical_status
            )

            if fill_color:
                value_cell.fill = PatternFill(
                    fill_type="solid",
                    fgColor=fill_color,
                )

            if font_color:
                value_cell.font = Font(
                    color=font_color,
                    bold=True,
                )

            break

    def _set_column_widths(
        self,
        worksheet,
        widths: list[int],
    ) -> None:
        for index, width in enumerate(
            widths,
            start=1,
        ):
            worksheet.column_dimensions[
                get_column_letter(index)
            ].width = width

    def _ratio_value(
        self,
        value: Any,
    ) -> float | None:
        if value is None:
            return None

        try:
            ratio = float(value)
        except (TypeError, ValueError):
            return None

        return min(
            max(ratio, 0.0),
            1.0,
        )

    def _format_json_value(
        self,
        value: Any,
    ) -> str:
        if value in (None, "", [], {}, ()):
            return ""

        if isinstance(value, str):
            return value

        try:
            return json.dumps(
                value,
                ensure_ascii=False,
                sort_keys=True,
            )
        except (TypeError, ValueError):
            return str(value)

    def _yes_no(
        self,
        value: Any,
    ) -> str:
        return "Sim" if bool(value) else "Não"

    def _set_hyperlink(
        self,
        cell,
    ) -> None:
        value = cell.value

        if value is None:
            return

        target = str(value).strip()

        if not target:
            return

        cell.hyperlink = target
        cell.style = "Hyperlink"

    def _format_datetime(
        self,
        value,
    ) -> str:
        if value is None:
            return "Não informada"

        return value.strftime(
            "%d/%m/%Y às %H:%M:%S"
        )

    def _format_confidence(
        self,
        value: Any,
    ) -> str:
        if value is None:
            return "Não informada"

        try:
            confidence = float(value)
        except (TypeError, ValueError):
            return str(value)

        confidence = min(
            max(confidence, 0.0),
            1.0,
        )

        return f"{confidence * 100:.0f}%"

    def _translate_severity(
        self,
        severity: EvidenceSeverity,
    ) -> str:
        labels = {
            EvidenceSeverity.CRITICAL: "Crítico",
            EvidenceSeverity.HIGH: "Alto",
            EvidenceSeverity.MEDIUM: "Médio",
            EvidenceSeverity.LOW: "Baixo",
            EvidenceSeverity.INFO: "Informativo",
        }

        return labels.get(
            severity,
            "Não classificado",
        )

    def _translate_severity_value(
        self,
        severity: Any,
    ) -> str:
        normalized_severity = str(
            severity or ""
        ).strip().lower()

        labels = {
            "critical": "Crítico",
            "high": "Alto",
            "medium": "Médio",
            "low": "Baixo",
            "info": "Informativo",
        }

        return labels.get(
            normalized_severity,
            normalized_severity or "Não classificado",
        )

    def _translate_batch_status(
        self,
        status: str,
    ) -> str:
        labels = {
            "pending": "Aguardando processamento",
            "processing": "Em processamento",
            "completed": "Concluído",
            "completed_with_errors": (
                "Concluído com ocorrências"
            ),
            "failed": "Falhou",
        }

        return labels.get(
            status,
            status,
        )

    def _translate_document_status(
        self,
        status: str,
    ) -> str:
        labels = {
            "pending": "Aguardando",
            "processing": "Em processamento",
            "completed": "Concluído",
            "failed": "Falhou",
        }

        return labels.get(
            status,
            status,
        )

    def _optional_uuid(
        self,
        value,
    ) -> str | None:
        if value is None:
            return None

        return str(value)

    def _safe_attribute(
        self,
        value: Any,
        attribute: str,
    ) -> Any:
        if value is None:
            return None

        if isinstance(value, dict):
            return value.get(
                attribute
            )

        return getattr(
            value,
            attribute,
            None,
        )

    def _enum_value(
        self,
        value: Any,
    ) -> Any:
        if value is None:
            return None

        return getattr(
            value,
            "value",
            value,
        )

    def _join_values(
        self,
        values: list[Any],
    ) -> str:
        normalized_values: list[str] = []

        for value in values:
            if value is None:
                continue

            normalized_value = str(
                value
            ).strip()

            if (
                not normalized_value
                or normalized_value
                in normalized_values
            ):
                continue

            normalized_values.append(
                normalized_value
            )

        return "\n".join(
            normalized_values
        )
