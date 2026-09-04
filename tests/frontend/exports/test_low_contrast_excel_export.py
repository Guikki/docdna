from __future__ import annotations

from openpyxl import Workbook

from app.domain.services.batch_excel_export_service import (
    BatchExcelExportService,
)
from app.frontend.exports.batch_export_view_builder import (
    BatchExportViewBuilder,
)


def _analysis_view() -> dict:
    white_location = {
        "finding_index": 1,
        "finding_code": "near_white_text",
        "detector": "white_text_detector",
        "page_number": 1,
        "matched_content": "texto branco",
        "left": 10.0,
        "top": 20.0,
        "width": 80.0,
        "height": 12.0,
        "confidence": 0.82,
        "confidence_label": "82.0%",
        "font_name": "Helvetica",
        "font_size": "8.00",
        "font_color_hex": "#FFFFFF",
        "coordinates_label": (
            "X: 10.00, Y: 20.00, "
            "largura: 80.00, altura: 12.00"
        ),
        "located": True,
        "message": "Região localizada.",
        "source_image_url": (
            "/reports/white-source.png"
        ),
        "annotated_image_url": (
            "/reports/white-annotated.png"
        ),
    }

    low_contrast_location = {
        "finding_index": 2,
        "finding_code": "low_contrast_text",
        "detector": "low_contrast_text_detector",
        "page_number": 1,
        "matched_content": "texto cinza",
        "left": 100.0,
        "top": 120.0,
        "width": 90.0,
        "height": 14.0,
        "confidence": 0.94,
        "confidence_label": "94.0%",
        "font_name": "Helvetica",
        "font_size": "9.00",
        "font_color_hex": "#BBBBBB",
        "coordinates_label": (
            "X: 100.00, Y: 120.00, "
            "largura: 90.00, altura: 14.00"
        ),
        "located": True,
        "message": "Região localizada.",
        "source_image_url": (
            "/reports/low-source.png"
        ),
        "annotated_image_url": (
            "/reports/low-annotated.png"
        ),
    }

    return {
        "id": "analysis-1",
        "filename": "documento.pdf",
        "visual_concealment_white_text_findings": [
            {
                "code": "near_white_text",
                "detector": "white_text_detector",
                "page_number": 1,
                "text": "texto branco",
                "font_name": "Helvetica",
                "font_size": "8.00",
                "font_color_hex": "#FFFFFF",
                "confidence": 0.82,
                "confidence_label": "82.0%",
                "signals": [
                    "near_white_font"
                ],
                "signal_labels": [
                    "Fonte branca ou quase branca"
                ],
                "is_near_white": True,
                "is_small_text": False,
                "is_relative_small_text": False,
                "is_instruction_like": False,
                "coordinates_label": (
                    "X: 10.00, Y: 20.00, "
                    "largura: 80.00, altura: 12.00"
                ),
                "visual_location": white_location,
            }
        ],
        "visual_concealment_low_contrast_text_findings": [
            {
                "code": "low_contrast_text",
                "detector": "low_contrast_text_detector",
                "page_number": 1,
                "text": "texto cinza",
                "font_name": "Helvetica",
                "font_size": "9.00",
                "font_color_hex": "#BBBBBB",
                "background_color_hex": "#FFFFFF",
                "font_relative_luminance": 0.497,
                "background_relative_luminance": 1.0,
                "contrast_ratio": 1.92,
                "contrast_threshold": 2.0,
                "contrast_level": "low_contrast",
                "contrast_level_label": (
                    "Contraste baixo"
                ),
                "background_dominance_ratio": 0.91,
                "background_sampling_method": (
                    "dominant_quantized_bbox_color"
                ),
                "background_sampling_method_label": (
                    "Cor dominante quantizada na BoundingBox"
                ),
                "contrast_reference": (
                    "Razão de contraste por luminância "
                    "relativa WCAG"
                ),
                "confidence": 0.94,
                "confidence_label": "94.0%",
                "signals": [
                    "low_contrast",
                    "background_color_estimated",
                ],
                "signal_labels": [
                    "Baixo contraste entre texto e fundo",
                    "Cor de fundo estimada localmente",
                ],
                "is_near_white": False,
                "is_low_contrast": True,
                "is_extreme_low_contrast": False,
                "is_small_text": False,
                "is_relative_small_text": False,
                "is_instruction_like": False,
                "coordinates_label": (
                    "X: 100.00, Y: 120.00, "
                    "largura: 90.00, altura: 14.00"
                ),
                "visual_location": low_contrast_location,
            }
        ],
        "visual_concealment_tiny_text_evidences": [
            {
                "code": "PROMPT_INJECTION_TINY_TEXT",
                "detector": "tiny_text",
                "page_number": 1,
                "text": "texto pequeno",
                "description": "Fonte abaixo do limite.",
                "font_name": "Helvetica",
                "font_size": "3.00",
                "font_color_hex": "#000000",
                "confidence": 0.70,
                "confidence_label": "70.0%",
                "coordinates_label": None,
            }
        ],
        "numeric_line_locations": [],
        "prompt_injection_locations": [],
        "visual_concealment_locations": [
            white_location,
            low_contrast_location,
        ],
    }


def test_export_builder_should_include_low_contrast_data() -> None:
    builder = BatchExportViewBuilder()

    records = builder._build_concealment(
        [_analysis_view()]
    )

    assert [
        record["concealment_type"]
        for record in records
    ] == [
        "white_text",
        "low_contrast_text",
        "tiny_text",
    ]

    low_contrast = records[1]

    assert low_contrast["finding_index"] == 2
    assert low_contrast["background_color_hex"] == "#FFFFFF"
    assert low_contrast["contrast_ratio"] == 1.92
    assert low_contrast["contrast_threshold"] == 2.0
    assert low_contrast["contrast_level"] == "low_contrast"
    assert low_contrast["background_dominance_ratio"] == 0.91
    assert low_contrast["is_low_contrast"] is True
    assert low_contrast["is_extreme_low_contrast"] is False
    assert low_contrast["located"] is True
    assert (
        low_contrast["annotated_image_url"]
        == "/reports/low-annotated.png"
    )


def test_locations_should_receive_low_contrast_context() -> None:
    builder = BatchExportViewBuilder()

    locations = builder._build_locations(
        [_analysis_view()]
    )

    assert len(locations) == 2

    low_location = next(
        location
        for location in locations
        if location["reference_code"]
        == "low_contrast_text"
    )

    assert low_location["reference_index"] == 2
    assert low_location["background_color_hex"] == "#FFFFFF"
    assert low_location["contrast_ratio"] == 1.92
    assert low_location["contrast_threshold"] == 2.0
    assert low_location["contrast_level"] == "low_contrast"
    assert low_location["background_dominance_ratio"] == 0.91


def test_concealment_sheet_should_export_contrast_columns() -> None:
    builder = BatchExportViewBuilder()
    record = builder._build_concealment(
        [_analysis_view()]
    )[1]

    service = object.__new__(
        BatchExcelExportService
    )

    workbook = Workbook()
    worksheet = workbook.active

    exported = service._build_concealment_sheet(
        worksheet=worksheet,
        export_view={
            "concealment": [record]
        },
    )

    assert exported == 1

    headers = {
        cell.value: index
        for index, cell in enumerate(
            worksheet[1],
            start=1,
        )
    }

    row = 2

    assert worksheet.cell(
        row=row,
        column=headers["Cor estimada do fundo"],
    ).value == "#FFFFFF"

    ratio_cell = worksheet.cell(
        row=row,
        column=headers["Razão de contraste"],
    )

    assert ratio_cell.value == 1.92
    assert ratio_cell.number_format == '0.00":1"'

    threshold_cell = worksheet.cell(
        row=row,
        column=headers["Limiar de contraste"],
    )

    assert threshold_cell.value == 2.0
    assert threshold_cell.number_format == '0.00":1"'

    dominance_cell = worksheet.cell(
        row=row,
        column=headers["Predominância do fundo"],
    )

    assert dominance_cell.value == 0.91
    assert dominance_cell.number_format == "0.0%"

    assert worksheet.cell(
        row=row,
        column=headers["Baixo contraste"],
    ).value == "Sim"

    annotated_cell = worksheet.cell(
        row=row,
        column=headers["Imagem anotada"],
    )

    assert (
        annotated_cell.hyperlink.target
        == "/reports/low-annotated.png"
    )


def test_locations_sheet_should_export_contrast_context() -> None:
    builder = BatchExportViewBuilder()

    low_location = next(
        location
        for location in builder._build_locations(
            [_analysis_view()]
        )
        if location["reference_code"]
        == "low_contrast_text"
    )

    service = object.__new__(
        BatchExcelExportService
    )

    workbook = Workbook()
    worksheet = workbook.active

    exported = service._build_locations_sheet(
        worksheet=worksheet,
        export_view={
            "locations": [low_location]
        },
    )

    assert exported == 1

    headers = {
        cell.value: index
        for index, cell in enumerate(
            worksheet[1],
            start=1,
        )
    }

    row = 2

    assert worksheet.cell(
        row=row,
        column=headers["Cor estimada do fundo"],
    ).value == "#FFFFFF"

    ratio_cell = worksheet.cell(
        row=row,
        column=headers["Razão de contraste"],
    )

    assert ratio_cell.value == 1.92
    assert ratio_cell.number_format == '0.00":1"'

    dominance_cell = worksheet.cell(
        row=row,
        column=headers["Predominância do fundo"],
    )

    assert dominance_cell.value == 0.91
    assert dominance_cell.number_format == "0.0%"

    annotated_cell = worksheet.cell(
        row=row,
        column=headers["Imagem anotada"],
    )

    assert (
        annotated_cell.hyperlink.target
        == "/reports/low-annotated.png"
    )
